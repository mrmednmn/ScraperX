import sys
sys.dont_write_bytecode = True
from saving_strategies.base import SavingStrategy
from saving_strategies.excel.header import ExcelHeader
import openpyxl
import os

class ExcelSavingStrategy(SavingStrategy):
    def __init__(self,folder_name:str,file_name:str, logger):
        self.logger=logger
        self.name="excel"
        self.create_folder(self.name)
        self.path_head = os.path.join(os.getcwd(),self.name,folder_name)
        self.create_folder(self.path_head)
        self.file_name=os.path.join(self.path_head, file_name+".xlsx")
        self.notify("Save Location: " + self.file_name)
        self.workbook=None
        self.sheet=None
        self.header = ExcelHeader()
        self.NA="N/A"

    #override
    def get_name(self):
        return self.name

    #override
    def notify(self,data):
        self.logger(data)

    #save the final file
    def save_file(self):
        try:
            self.workbook.save(self.file_name)
        except Exception as ex:
            self.notify("save_file " + str(ex))

    #init ouedkniss cars data excel file head.
    def init_cols_and_headings(self):
        #Open an existing workbook or create a new one if it doesn't exist
        self.workbook = openpyxl.load_workbook(self.file_name) if os.path.exists(self.file_name) else openpyxl.Workbook()
        #Select the active sheet or create a new one if it doesn't exist
        self.sheet = self.workbook.active if self.workbook.sheetnames else self.workbook.create_sheet()
        for column, heading in zip(self.header.columns, self.header.headings):
            self.sheet.column_dimensions[column].width = 50 
            self.sheet[f"{column}1"] = heading
        self.save_file()

    def build_columns(self,headings):
        for col_num in range(1, len(headings) + 1):
            column_letter = openpyxl.utils.get_column_letter(col_num)
            self.header.columns.append(column_letter)

    def init_excel(self,headings):
        self.build_columns(headings)
        self.header.set_headings(headings)
        self.init_cols_and_headings()

    def create_folder(self, folder_name: str):
        folder_path = os.path.join(os.getcwd(), folder_name)
        if not os.path.isdir(folder_path):
            os.makedirs(folder_path)

    def log(self,msg):
        print(msg,end='\n')

    #def clear_data(self,data):
        #for k in data:
            #data[k]=self.NA
    
    #get the active workbook
    def set_the_active_workbook(self):
        self.workbook=openpyxl.load_workbook(self.file_name)
        
    #get the active sheet
    def set_the_active_sheet(self):
        self.set_the_active_workbook()
        self.sheet= self.workbook.active

    def add_row(self, data, donot_repeat):
        title_index = self.header.headings.index(donot_repeat)
        title_value = data.get(donot_repeat)

        rows = self.sheet.iter_rows(min_row=2, values_only=True)
        matching_row = any(row[title_index] == title_value for row in rows)
    
        if matching_row:
            self.notify("row already exists, ignore..")
            return
    
        next_row = self.sheet.max_row + 1
        for column_index, column in enumerate(self.header.columns):
            cell = f"{column}{next_row}"
            value = data.get(self.header.headings[column_index], "")
            self.sheet[cell] = value
        
    #override
    def save(self,data,donot_repeat):
        data = {key: data[key] for key in sorted(data)}
        self.set_the_active_sheet()
        self.add_row(data,donot_repeat)
        #self.clear_data(data)
        self.save_file()
